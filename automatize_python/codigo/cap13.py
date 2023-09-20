#TRABAJAR CON HOJAS DE CALCULO CON EXEL
from pathlib import Path
import openpyxl
#para usar este modulo ya debe existir el documento de exel, y le pasamos la ruta de su ubicacion a la funcion
p = Path.home()/'desktop'
wb = openpyxl.load_workbook(p/'ejemplo.xlsx')
print(type(wb)) #<class 'openpyxl.workbook.workbook.Workbook'> - devuelve un objeto libro de trabajo

#1......obtener lista de hojas del libro de trabajo
lh = wb.sheetnames
print(lh) #['Hoja1', 'detalle', 'Hoja2']

#2......obetener una hoja del libro de trabajo
h = wb['Hoja1']
print(h) #<Worksheet "Hoja1">
print(type(h)) #<class 'openpyxl.worksheet.worksheet.Worksheet'> - devuelve un objeto hoja de trabajo(cada hoja representa un objeto hoja de trabajo)

#3......obtener el titulo de una hoja como cadena
ht = h.title
print(ht) #Hoja1

#4......obtener la hoja actualmente activa
ha = wb.active
print(ha) #<Worksheet "Hoja1"> - esta es la hoja donde se hizo la ultima actualizacion 'guardar' aunque el documento este cerrado

#5......obtener celdas de las hojas
wb1 = openpyxl.load_workbook(p/'ejemplo.xlsx')
h1 = wb1['Hoja1'] #este devuelve un objeto de hoja
h2= h1['B3'] #este devuelve un objeto celda que tiene el atributo valor (value), columna (column), coordenada (coordenate), fila
print(h2) #<Cell 'Hoja1'.B3>

#6......obtener valores de las celdas
vc = h1['B3'].value
print(vc) # 2015-04-05 13:34:02
vc1 = h1['C3'].value
print(vc1) #manzanas

#7......obtiene la fila, columna, valor y coordenada utilizando el objeto celda
c = 'fila es %s, columna es %s, valor es %s, coordenada es %s' %(h2.row, h2.column, h2.value, h2.coordinate)
print(c) #fila es 3, columna es 2, valor es 2015-04-05 13:34:02, coordenada es B3

#8......obtener una celda con metodo cell()
#al metodo cell() se le pasan numeros enteros para sus argumentos fila (row) y columna (column)
#para obtener la celda le pasamos el objeto hoja de trabajo 
c = h1.cell(row=2, column=2)
print(c) #<Cell 'Hoja1'.B2>

v = c.value
print(v) #A

for i in range (1,8,2):
    print(i, h1.cell(row=i, column=2).value)
'''
1 None
3 2015-04-05 13:34:02    
5 6/04/2015 12:46:51 p.m.
7 2015-04-10 02:07:00
'''  

#9......determinar el tamaño de la hoja con atributos maxima fila (max_row) y maxima columna (max_column)
#ambos atributos buscan la ultima fila y columna con datos existentes guardados
mr = h1.max_row
print(mr) #11
mr = h1.max_column
print(mr) #6

#10......conversion entre letras de colunmas y numeros
#para convertir de letras a numeros con la funcion (openpyxl.utils.column_index_from_string())
#primero debemos importar las funciones (from openpyxl.utils import get_column_letter, column_index_from_string) del modulo (openpyxl.utils)
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1)) #A
print(get_column_letter(2)) #B
print(get_column_letter(27)) #AA
print(get_column_letter(900)) #AHP
print(get_column_letter(h1.max_column)) #F

print(column_index_from_string('A')) #1
print(column_index_from_string('AA')) #27

